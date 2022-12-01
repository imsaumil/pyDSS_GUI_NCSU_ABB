# Importing required libraries
import openpyxl.drawing.image
import os
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
import glob

# Defining file path
file_path = os.path.dirname(os.path.abspath(__file__))

report_path = os.path.join(file_path, "reports", 'Simulation_Report.xlsx')


def xlsx_report_writer(sim_run_data_1, sim_run_data_2):
    pd.DataFrame.reset_header = lambda df: df.swapaxes(0, 1).reset_index().swapaxes(0, 1)

    # Reset header to include in xlsx
    sim_run_data_1 = sim_run_data_1.reset_header()
    sim_run_data_2 = sim_run_data_2.reset_header()

    # Starting to write an Excel sheet
    workbook = xlsxwriter.Workbook(report_path)

    # Setting sheet format
    bold = workbook.add_format({'bold': True, 'border': 1, 'align': 'center'})
    merge_format = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    default = workbook.add_format({'border': 1, 'align': 'center'})

    worksheet = workbook.add_worksheet('Report')
    worksheet.insert_image(0, 0, os.path.join(file_path, 'assets', 'sponsor_logo.png'))

    # Table start values
    start_row = 5
    start_col = 17

    for col in range(start_col, start_col + len(sim_run_data_1.columns)):
        worksheet.set_column(col, col, 18)

    worksheet.merge_range(start_row - 1, start_col, start_row - 1,
                          start_col + len(sim_run_data_1.columns) - 1, 'SIMULATION RESULTS', merge_format)

    # Writing and formatting table 1
    for i in range(len(sim_run_data_1)):
        for j in range(len(sim_run_data_1.columns)):
            if i == 0:
                worksheet.write(i + start_row, j + start_col, sim_run_data_1.iloc[i, j], bold)
            else:
                worksheet.write(i + start_row, j + start_col, sim_run_data_1.iloc[i, j], default)

    # Specifying space between two tables
    mid_row = start_row + len(sim_run_data_1) + 4

    worksheet.merge_range(mid_row - 1, start_col, mid_row - 1,
                          start_col + len(sim_run_data_2.columns) - 1, 'FAULT CONTRIBUTION RESULTS', merge_format)

    # Writing and formatting table 2
    for i in range(len(sim_run_data_2)):
        for j in range(len(sim_run_data_2.columns)):
            if i == 0:
                worksheet.write(i + mid_row, j + start_col, sim_run_data_2.iloc[i, j], bold)
            else:
                worksheet.write(i + mid_row, j + start_col, sim_run_data_2.iloc[i, j], default)

    workbook.close()


def image_insert():

    # Loading the previously created file to just add an image
    workbook = load_workbook(report_path)
    ws = workbook.active

    download_path = os.path.join(os.path.expanduser("~"), 'Downloads')
    img_path = os.path.join(download_path, 'ieee123_bus.jpg')

    if os.path.isfile(img_path):
        img = openpyxl.drawing.image.Image(img_path)
    else:
        img = max(glob.glob(download_path + r"\*jpg"), key=os.path.getctime)

    img.anchor = 'A5'
    ws.add_image(img)

    workbook.save(os.path.join(report_path))

    os.remove(os.path.join(download_path, 'ieee123_bus.jpg'))
